"""export_service — Excel workbook export (Phase 4.5, Epic E4, Issue #4.1).

Builds a single ``.xlsx`` workbook with three sheets — **Tài sản**, **Thu
chi**, **Mục tiêu** — so the user can walk away with their whole picture in
a file they own. This is the "data portability" promise of a *người đồng
hành quản lý tài sản*: their numbers are theirs to take.

Design — pure builder + thin I/O shell (mirrors ``clarity_service``):
    ``gather_export_data(db, user_id)`` runs a handful of indexed, read-only
    queries and packs the rows into :class:`ExportData`.
    ``build_workbook(data)`` is a **pure** function (no DB, no clock, no LLM)
    that turns those rows into workbook ``bytes``. Splitting the two keeps
    the layout trivially unit-testable (round-trip a fabricated
    :class:`ExportData` with no database) and keeps the I/O contract honest.

Money contract: every amount stays a :class:`~decimal.Decimal` all the way
into the cell — ``openpyxl`` writes ``Decimal`` as a native number so no
precision is lost to ``float`` (CLAUDE.md §Money handling). Cells are given
a thousands number format so the sheet reads "1,500,000", but the stored
value is exact.

Layer contract: this is a service — it is **read-only** (no writes, no
flush, no commit), never sends Telegram, and never reads env. The
``EXPORT_EXCEL_ENABLED`` flag is read by the handler/router edge (see the
worker), never here. User-facing strings (headers, labels) come from
``content/export_copy.yaml`` — never hardcoded.
"""

from __future__ import annotations

import io
import uuid
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal
from functools import lru_cache
from pathlib import Path

import yaml
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from backend.models.expense import Expense
from backend.models.goal import Goal
from backend.models.income_record import IncomeRecord
from backend.wealth.models.asset import Asset

_COPY_PATH = Path(__file__).resolve().parents[3] / "content" / "export_copy.yaml"

# openpyxl number format for VND amounts — thousands separator, no decimals
# (VND has no minor unit in practice). The underlying cell value stays an
# exact Decimal; this only controls display.
_MONEY_FMT = "#,##0"
_DATE_FMT = "yyyy-mm-dd"


@lru_cache(maxsize=1)
def _copy() -> dict:
    with _COPY_PATH.open("r", encoding="utf-8") as fh:
        return yaml.safe_load(fh) or {}


def _sheet_copy(key: str) -> dict:
    return dict((_copy().get("sheets") or {}).get(key) or {})


def _asset_type_label(asset_type: str) -> str:
    labels = _copy().get("asset_type_labels") or {}
    return str(labels.get(asset_type, asset_type))


# ---------------------------------------------------------------------------
# Data structures — plain rows, decoupled from ORM objects so the pure
# builder never touches a detached SQLAlchemy instance.
# ---------------------------------------------------------------------------


@dataclass(frozen=True, slots=True)
class AssetRow:
    asset_type: str
    name: str
    current_value: Decimal
    initial_value: Decimal
    gain_loss: Decimal
    acquired_at: date | None
    last_valued_at: datetime | None


@dataclass(frozen=True, slots=True)
class CashflowRow:
    is_income: bool
    on_date: date | None
    label: str  # category (expense) or source (income)
    note: str
    amount: Decimal


@dataclass(frozen=True, slots=True)
class GoalRow:
    name: str
    target_amount: Decimal
    current_amount: Decimal
    remaining_amount: Decimal
    progress_pct: Decimal
    target_date: date | None


@dataclass(frozen=True, slots=True)
class ExportData:
    """Everything the workbook needs, gathered once, ORM-free."""

    assets: tuple[AssetRow, ...] = field(default_factory=tuple)
    cashflow: tuple[CashflowRow, ...] = field(default_factory=tuple)
    goals: tuple[GoalRow, ...] = field(default_factory=tuple)

    @property
    def is_empty(self) -> bool:
        return not (self.assets or self.cashflow or self.goals)


# ---------------------------------------------------------------------------
# Pure builder — deterministic, no I/O.
# ---------------------------------------------------------------------------


def _write_header(ws: Worksheet, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)


def _money_cell(ws: Worksheet, row: int, col: int, value: Decimal) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.number_format = _MONEY_FMT


def _build_assets_sheet(ws: Worksheet, data: ExportData) -> None:
    copy = _sheet_copy("assets")
    ws.title = copy.get("title", "Tài sản")
    _write_header(ws, list(copy.get("headers") or []))
    for a in data.assets:
        r = ws.max_row + 1
        ws.cell(row=r, column=1, value=_asset_type_label(a.asset_type))
        ws.cell(row=r, column=2, value=a.name)
        _money_cell(ws, r, 3, a.current_value)
        _money_cell(ws, r, 4, a.initial_value)
        _money_cell(ws, r, 5, a.gain_loss)
        ws.cell(row=r, column=6, value=a.acquired_at).number_format = _DATE_FMT
        ws.cell(
            row=r, column=7, value=_naive(a.last_valued_at)
        ).number_format = _DATE_FMT


def _build_cashflow_sheet(ws: Worksheet, data: ExportData) -> None:
    copy = _sheet_copy("cashflow")
    ws.title = copy.get("title", "Thu chi")
    _write_header(ws, list(copy.get("headers") or []))
    income_label = copy.get("row_income", "Thu")
    expense_label = copy.get("row_expense", "Chi")
    for c in data.cashflow:
        r = ws.max_row + 1
        ws.cell(
            row=r, column=1, value=income_label if c.is_income else expense_label
        )
        ws.cell(row=r, column=2, value=c.on_date).number_format = _DATE_FMT
        ws.cell(row=r, column=3, value=c.label)
        ws.cell(row=r, column=4, value=c.note)
        _money_cell(ws, r, 5, c.amount)


def _build_goals_sheet(ws: Worksheet, data: ExportData) -> None:
    copy = _sheet_copy("goals")
    ws.title = copy.get("title", "Mục tiêu")
    _write_header(ws, list(copy.get("headers") or []))
    for g in data.goals:
        r = ws.max_row + 1
        ws.cell(row=r, column=1, value=g.name)
        _money_cell(ws, r, 2, g.target_amount)
        _money_cell(ws, r, 3, g.current_amount)
        _money_cell(ws, r, 4, g.remaining_amount)
        pct = ws.cell(row=r, column=5, value=g.progress_pct)
        pct.number_format = "0.0"
        ws.cell(row=r, column=6, value=g.target_date).number_format = _DATE_FMT


def _naive(dt: datetime | None) -> datetime | None:
    """openpyxl rejects tz-aware datetimes — drop tzinfo for the cell."""
    if dt is None:
        return None
    return dt.replace(tzinfo=None) if dt.tzinfo is not None else dt


def build_workbook(data: ExportData) -> bytes:
    """Turn gathered rows into a 3-sheet ``.xlsx`` as bytes.

    Pure and deterministic: the same :class:`ExportData` always yields an
    equivalent workbook. An empty ``data`` still produces all three sheets
    with their header rows (never crashes, never an empty file) so the user
    always gets a well-formed workbook back.
    """
    wb = Workbook()
    _build_assets_sheet(wb.active, data)
    _build_cashflow_sheet(wb.create_sheet(), data)
    _build_goals_sheet(wb.create_sheet(), data)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# I/O shell — gather rows then delegate to the pure builder.
# ---------------------------------------------------------------------------


async def gather_export_data(db: AsyncSession, user_id: uuid.UUID) -> ExportData:
    """Read the three datasets with lightweight indexed queries.

    Read-only: no writes, no flush, no commit. Mirrors the real-asset /
    non-deleted filters used elsewhere so demo/placeholder rows never leak
    into the export.
    """
    # Assets — real rows only (mirrors ``get_user_assets`` / clarity_service).
    asset_rows = (
        await db.execute(
            select(Asset)
            .where(
                Asset.user_id == user_id,
                Asset.is_active.is_(True),
                Asset.is_placeholder_asset.is_(False),
                Asset.is_confirmed.is_(True),
            )
            .order_by(Asset.current_value.desc())
        )
    ).scalars().all()
    assets = tuple(
        AssetRow(
            asset_type=a.asset_type,
            name=a.name,
            current_value=Decimal(a.current_value or 0),
            initial_value=Decimal(a.initial_value or 0),
            gain_loss=a.gain_loss,
            acquired_at=a.acquired_at,
            last_valued_at=a.last_valued_at,
        )
        for a in asset_rows
    )

    # Income — non-deleted, newest first.
    income_rows = (
        await db.execute(
            select(IncomeRecord)
            .where(
                IncomeRecord.user_id == user_id,
                IncomeRecord.deleted_at.is_(None),
            )
            .order_by(IncomeRecord.period.desc())
        )
    ).scalars().all()

    # Expenses — non-deleted, newest first.
    expense_rows = (
        await db.execute(
            select(Expense)
            .where(
                Expense.user_id == user_id,
                Expense.deleted_at.is_(None),
            )
            .order_by(Expense.expense_date.desc())
        )
    ).scalars().all()

    cashflow: list[CashflowRow] = []
    for inc in income_rows:
        cashflow.append(
            CashflowRow(
                is_income=True,
                on_date=inc.period,
                label=inc.source or inc.income_type or "",
                note=inc.note or "",
                amount=Decimal(inc.amount or 0),
            )
        )
    for exp in expense_rows:
        cashflow.append(
            CashflowRow(
                is_income=False,
                on_date=exp.expense_date,
                label=exp.category or exp.merchant or "",
                note=exp.note or exp.merchant or "",
                amount=Decimal(exp.amount or 0),
            )
        )
    # Interleave by date, newest first; rows without a date sort last.
    cashflow.sort(key=lambda c: (c.on_date is not None, c.on_date), reverse=True)

    # Goals — active, non-deleted.
    goal_rows = (
        await db.execute(
            select(Goal)
            .where(
                Goal.user_id == user_id,
                Goal.status == "active",
                Goal.deleted_at.is_(None),
            )
            .order_by(Goal.priority.asc(), Goal.created_at.asc())
        )
    ).scalars().all()
    goals = tuple(
        GoalRow(
            name=g.name,
            target_amount=Decimal(g.target_amount or 0),
            current_amount=Decimal(g.current_amount or 0),
            remaining_amount=g.remaining_amount,
            progress_pct=Decimal(str(round(g.progress_pct, 1))),
            target_date=g.target_date,
        )
        for g in goal_rows
    )

    return ExportData(assets=assets, cashflow=tuple(cashflow), goals=goals)


async def build_export(db: AsyncSession, user_id: uuid.UUID) -> tuple[bytes, bool]:
    """Gather + build in one call. Returns ``(xlsx_bytes, is_empty)``.

    ``is_empty`` lets the handler pick the right caption (a warm nudge to add
    data vs. the normal "here's your file"). Read-only, no LLM.
    """
    data = await gather_export_data(db, user_id)
    return build_workbook(data), data.is_empty
