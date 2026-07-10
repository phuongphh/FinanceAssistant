"""Phase 4.5 / E4 / Issue #4.1 — Excel export.

Two layers under test:

1. ``export_service`` — the pure ``build_workbook`` builder (round-trip via
   openpyxl, numbers land exactly, empty data still yields well-formed
   header-only sheets) plus ``gather_export_data`` row mapping (DB mocked, so
   these stay fast and DB-free).
2. ``export_handler.cmd_export`` — the flag-gated edge (flag dark → "tạm tắt"
   note and never builds; no user → gentle nudge; happy path ships a document
   through the Notifier port).

Money contract: amounts are Decimals all the way into the cell; openpyxl
writes them natively so no float precision is lost (CLAUDE.md §Money handling).
"""

from __future__ import annotations

import io
import uuid
from datetime import date, datetime, timezone
from decimal import Decimal
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook

from backend.services.export import export_service
from backend.services.export.export_service import (
    AssetRow,
    CashflowRow,
    ExportData,
    GoalRow,
    build_workbook,
)

_BANNED = ("Decision Engine", "CFO", "GPS")


# ---------------------------------------------------------------------------
# Fabricated data — mirrors what gather_export_data would produce.
# ---------------------------------------------------------------------------


def _export_data() -> ExportData:
    return ExportData(
        assets=(
            AssetRow(
                asset_type="cash",
                name="Tiền mặt",
                current_value=Decimal("150000000"),
                initial_value=Decimal("150000000"),
                gain_loss=Decimal("0"),
                acquired_at=date(2026, 1, 1),
                last_valued_at=datetime(2026, 7, 1, tzinfo=timezone.utc),
            ),
            AssetRow(
                asset_type="stock",
                name="HPG",
                current_value=Decimal("52000000"),
                initial_value=Decimal("40000000"),
                gain_loss=Decimal("12000000"),
                acquired_at=date(2025, 3, 15),
                last_valued_at=datetime(2026, 7, 5, tzinfo=timezone.utc),
            ),
        ),
        cashflow=(
            CashflowRow(
                is_income=True,
                on_date=date(2026, 7, 1),
                label="Lương",
                note="",
                amount=Decimal("30000000"),
            ),
            CashflowRow(
                is_income=False,
                on_date=date(2026, 6, 20),
                label="Ăn uống",
                note="cơm trưa",
                amount=Decimal("250000"),
            ),
        ),
        goals=(
            GoalRow(
                name="Quỹ khẩn cấp",
                target_amount=Decimal("100000000"),
                current_amount=Decimal("60000000"),
                remaining_amount=Decimal("40000000"),
                progress_pct=Decimal("60.0"),
                target_date=date(2027, 1, 1),
            ),
        ),
    )


def _load(xlsx: bytes):
    return load_workbook(io.BytesIO(xlsx))


# ---------------------------------------------------------------------------
# build_workbook — pure builder
# ---------------------------------------------------------------------------


def test_build_workbook_has_three_sheets():
    wb = _load(build_workbook(_export_data()))
    assert len(wb.sheetnames) == 3


def test_build_workbook_numbers_match_source_exactly():
    """Every money cell round-trips to the same value it was given —
    Decimal in, exact number out (no float drift)."""
    data = _export_data()
    wb = _load(build_workbook(data))

    assets_ws = wb.worksheets[0]
    # header row + 2 assets; column 3 = current_value.
    current_values = [assets_ws.cell(row=r, column=3).value for r in (2, 3)]
    assert [Decimal(str(v)) for v in current_values] == [
        Decimal("150000000"),
        Decimal("52000000"),
    ]
    # gain_loss (column 5) of the stock row must survive exactly.
    assert Decimal(str(assets_ws.cell(row=3, column=5).value)) == Decimal(
        "12000000"
    )

    goals_ws = wb.worksheets[2]
    assert Decimal(str(goals_ws.cell(row=2, column=2).value)) == Decimal(
        "100000000"
    )
    assert Decimal(str(goals_ws.cell(row=2, column=4).value)) == Decimal(
        "40000000"
    )


def test_build_workbook_headers_present_and_bold():
    wb = _load(build_workbook(_export_data()))
    for ws in wb.worksheets:
        header_cells = [c for c in ws[1] if c.value not in (None, "")]
        assert header_cells, f"sheet {ws.title!r} has no header"
        assert all(c.font.bold for c in header_cells)


def test_build_workbook_strips_tzinfo():
    """openpyxl rejects tz-aware datetimes — the builder must not crash and
    the stored datetime must be naive."""
    wb = _load(build_workbook(_export_data()))
    assets_ws = wb.worksheets[0]
    valued = assets_ws.cell(row=2, column=7).value
    assert isinstance(valued, datetime)
    assert valued.tzinfo is None


def test_build_workbook_empty_is_header_only_no_crash():
    """DoD: an empty user still gets a well-formed 3-sheet workbook with just
    the header rows — never an empty file, never a crash."""
    wb = _load(build_workbook(ExportData()))
    assert len(wb.sheetnames) == 3
    for ws in wb.worksheets:
        # exactly one non-empty row: the header.
        non_empty_rows = [
            r for r in ws.iter_rows() if any(c.value not in (None, "") for c in r)
        ]
        assert len(non_empty_rows) == 1


def test_export_data_is_empty_property():
    assert ExportData().is_empty is True
    assert _export_data().is_empty is False


# ---------------------------------------------------------------------------
# gather_export_data — row mapping (DB mocked)
# ---------------------------------------------------------------------------


class _Result:
    def __init__(self, rows):
        self._rows = rows

    def scalars(self):
        return self

    def all(self):
        return self._rows


class _FakeDB:
    """Returns queued row batches in call order: assets, income, expense,
    goals — matching gather_export_data's query sequence."""

    def __init__(self, batches):
        self._batches = list(batches)

    async def execute(self, *_a, **_k):
        return _Result(self._batches.pop(0))


@pytest.mark.asyncio
async def test_gather_maps_rows_and_sorts_cashflow_newest_first():
    asset = SimpleNamespace(
        asset_type="gold",
        name="SJC",
        current_value=Decimal("80000000"),
        initial_value=Decimal("70000000"),
        gain_loss=Decimal("10000000"),
        acquired_at=date(2025, 1, 1),
        last_valued_at=datetime(2026, 7, 1, tzinfo=timezone.utc),
    )
    income = SimpleNamespace(
        period=date(2026, 6, 1),
        source="Lương",
        income_type="salary",
        note=None,
        amount=Decimal("30000000"),
    )
    expense = SimpleNamespace(
        expense_date=date(2026, 7, 3),
        category="Ăn uống",
        merchant="Highlands",
        note=None,
        amount=Decimal("120000"),
    )
    goal = SimpleNamespace(
        name="Nhà",
        target_amount=Decimal("2000000000"),
        current_amount=Decimal("500000000"),
        remaining_amount=Decimal("1500000000"),
        progress_pct=25.0,
        target_date=date(2030, 1, 1),
    )
    db = _FakeDB([[asset], [income], [expense], [goal]])

    data = await export_service.gather_export_data(db, uuid.uuid4())

    assert len(data.assets) == 1 and data.assets[0].name == "SJC"
    assert data.assets[0].current_value == Decimal("80000000")
    # cashflow: 2 rows, newest (expense 07-03) first.
    assert [c.is_income for c in data.cashflow] == [False, True]
    assert data.cashflow[0].amount == Decimal("120000")
    assert len(data.goals) == 1
    assert data.goals[0].progress_pct == Decimal("25.0")


@pytest.mark.asyncio
async def test_build_export_empty_user_reports_empty():
    db = _FakeDB([[], [], [], []])
    xlsx, is_empty = await export_service.build_export(db, uuid.uuid4())
    assert is_empty is True
    # still a valid, openable workbook.
    assert len(_load(xlsx).sheetnames) == 3


# ---------------------------------------------------------------------------
# Copy hygiene
# ---------------------------------------------------------------------------


def test_export_copy_has_no_banned_positioning_words():
    blob = "\n".join(str(v) for v in _flatten(export_service._copy()))
    for banned in _BANNED:
        assert banned not in blob


def _flatten(obj):
    if isinstance(obj, dict):
        for v in obj.values():
            yield from _flatten(v)
    elif isinstance(obj, (list, tuple)):
        for v in obj:
            yield from _flatten(v)
    else:
        yield obj


# ---------------------------------------------------------------------------
# cmd_export — flag-gated edge
# ---------------------------------------------------------------------------


class _FakeNotifier:
    def __init__(self):
        self.messages: list[str] = []
        self.documents: list[tuple] = []

    async def send_message(self, chat_id, text, **_k):
        self.messages.append(text)
        return {"ok": True}

    async def send_document(self, chat_id, document, filename, **_k):
        self.documents.append((filename, document))
        return {"ok": True}


def _install_notifier(monkeypatch):
    from backend.bot.handlers import export_handler

    fake = _FakeNotifier()
    monkeypatch.setattr(export_handler, "get_notifier", lambda: fake)
    return fake, export_handler


@pytest.mark.asyncio
async def test_cmd_export_flag_off_sends_disabled_never_builds(monkeypatch):
    monkeypatch.setenv("EXPORT_EXCEL_ENABLED", "false")
    fake, export_handler = _install_notifier(monkeypatch)

    async def _boom(*_a, **_k):
        raise AssertionError("build_export must not run when flag is off")

    monkeypatch.setattr(export_service, "build_export", _boom)

    user = SimpleNamespace(id=uuid.uuid4(), salutation="anh")
    await export_handler.cmd_export(db=None, chat_id=1, user=user)

    assert fake.documents == []
    assert len(fake.messages) == 1  # the "tạm tắt" note only


@pytest.mark.asyncio
async def test_cmd_export_no_user_sends_nudge(monkeypatch):
    monkeypatch.setenv("EXPORT_EXCEL_ENABLED", "true")
    fake, export_handler = _install_notifier(monkeypatch)

    await export_handler.cmd_export(db=None, chat_id=1, user=None)

    assert fake.documents == []
    assert len(fake.messages) == 1  # not_registered nudge


@pytest.mark.asyncio
async def test_cmd_export_happy_path_ships_document(monkeypatch):
    monkeypatch.setenv("EXPORT_EXCEL_ENABLED", "true")
    fake, export_handler = _install_notifier(monkeypatch)

    async def _fake_build(db, user_id):
        return build_workbook(_export_data()), False

    monkeypatch.setattr(export_service, "build_export", _fake_build)

    user = SimpleNamespace(id=uuid.uuid4(), salutation="chị")
    await export_handler.cmd_export(db=None, chat_id=99, user=user)

    assert len(fake.documents) == 1
    filename, blob = fake.documents[0]
    assert filename.endswith(".xlsx")
    # a real, openable workbook came through the port.
    assert len(_load(blob).sheetnames) == 3
